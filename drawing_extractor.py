"""
Drawing Extractor Module
Excelファイル内の図形、SmartArt、画像などの描画要素を抽出するモジュール
"""

import os
import xml.etree.ElementTree as ET
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.utils import get_column_letter, column_index_from_string
from logger import Logger
import zipfile
import base64

class DrawingExtractor:

    def __init__(self):
        self.logger = Logger()

    def extract_drawing_info(self, excel_zip, sheet_name, drawing_path):
        """シートから描画要素(図形、SmartArt、画像など)の情報を抽出"""
        try:
            if drawing_path not in excel_zip.namelist():
                return []

            with excel_zip.open(drawing_path) as f:
                tree = ET.parse(f)
                root = tree.getroot()

                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'dgm': 'http://schemas.openxmlformats.org/drawingml/2006/diagram'
                }

                drawing_info = []
                for anchor in root.findall('./xdr:twoCellAnchor', ns) + root.findall('./xdr:oneCellAnchor', ns) + root.findall('./xdr:absoluteAnchor', ns):
                    drawing = self._process_anchor(anchor, ns, excel_zip, drawing_path)
                    if drawing:
                        drawing_info.append(drawing)

                return drawing_info

        except Exception as e:
            self.logger.error(f"Error extracting drawing info: {str(e)}")
            return []

    def _process_anchor(self, anchor, ns, excel_zip, drawing_path):
        """アンカー要素の処理"""
        try:
            from_cell = anchor.find('.//xdr:from', ns)
            to_cell = anchor.find('.//xdr:to', ns)

            if from_cell is None:
                return None

            from_row = int(from_cell.find('xdr:row', ns).text) + 1
            from_col = int(from_cell.find('xdr:col', ns).text) + 1

            to_row = from_row
            to_col = from_col

            if to_cell is not None:
                to_row = int(to_cell.find('xdr:row', ns).text) + 1
                to_col = int(to_cell.find('xdr:col', ns).text) + 1

            range_str = f"{get_column_letter(from_col)}{from_row}:{get_column_letter(to_col)}{to_row}"

            # SmartArt
            smartart = anchor.find('.//dgm:relIds', ns)
            if smartart is not None:
                smartart_info = self._extract_smartart_info(smartart, excel_zip, drawing_path, range_str)
                if smartart_info:
                    return smartart_info

            # Shape
            sp = anchor.find('.//xdr:sp', ns)
            if sp is not None:
                shape_info = self._extract_shape_info(sp, ns, range_str)
                if shape_info:
                    return shape_info

            # Image
            pic = anchor.find('.//xdr:pic', ns)
            if pic is not None:
                image_info = self.extract_picture_info(pic, excel_zip, ns, drawing_path)
                if image_info:
                    image_info["range"] = range_str
                    return image_info

            return None

        except Exception as e:
            self.logger.error(f"Error processing anchor: {str(e)}")
            return None

    def _extract_smartart_info(self, smartart, excel_zip, drawing_path, range_str):
        """SmartArt情報の抽出"""
        try:
            diagram_data = None
            data_model_rel_id = smartart.get('dm')
            if data_model_rel_id:
                diagram_data = self._extract_diagram_data(excel_zip, data_model_rel_id, drawing_path)

            if diagram_data:
                return {
                    "regionType": "smartart",
                    "range": range_str,
                    **diagram_data
                }

            return None

        except Exception as e:
            self.logger.error(f"Error extracting SmartArt info: {str(e)}")
            return None

    def _extract_diagram_data(self, excel_zip, rel_id, drawing_path):
        """SmartArtのダイアグラムデータを抽出"""
        try:
            drawing_number = os.path.basename(drawing_path).replace('drawing', '').replace('.xml', '')
            rels_path = f'xl/drawings/_rels/drawing{drawing_number}.xml.rels'

            diagram_path = None
            if rels_path in excel_zip.namelist():
                with excel_zip.open(rels_path) as rels_file:
                    rels_tree = ET.parse(rels_file)
                    rels_root = rels_tree.getroot()

                    for rel in rels_root.findall(
                        './/{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'
                    ):
                        if rel.get('Id') == rel_id:
                            target = rel.get('Target').replace('..', 'xl')
                            if not target.startswith('xl/'):
                                target = 'xl/' + target
                            diagram_path = target
                            break

            if not diagram_path or diagram_path not in excel_zip.namelist():
                self.logger.debug("SmartArt(ダイアグラム)に相当するファイルが見つかりませんでした。")
                return None

            with excel_zip.open(diagram_path) as f:
                tree = ET.parse(f)
                root = tree.getroot()

                ns = {
                    'dgm': 'http://schemas.openxmlformats.org/drawingml/2006/diagram',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                }

                diagram_data = {
                    "diagram_type": root.get('type', ''),
                    "name": root.get('name', ''),
                    "description": root.get('description', ''),
                    "diagram_file": diagram_path,
                    "nodes": []
                }

                nodes = root.findall('.//dgm:pt', ns)
                for node in nodes:
                    node_id = node.get('modelId')
                    all_a_t_elems = node.findall('.//a:t', ns)
                    texts = [el.text for el in all_a_t_elems if el.text]
                    diagram_data['nodes'].append({
                        'id': node_id,
                        'text_list': texts,
                    })

                return dict(diagram_data)

        except Exception as e:
            self.logger.error(f"Error extracting diagram data: {str(e)}")
            return None

    def _extract_shape_info(self, sp, ns, range_str):
        """図形情報の抽出"""
        try:
            shape_properties = sp.find('.//xdr:spPr', ns)
            shape_type = None
            if shape_properties is not None:
                prstGeom = shape_properties.find('.//a:prstGeom', ns)
                if prstGeom is not None:
                    shape_type = prstGeom.get('prst')

            text_body = sp.find('.//xdr:txBody', ns)
            text_content = None
            if text_body is not None:
                paragraphs = text_body.findall('.//a:p', ns)
                texts = []
                for p in paragraphs:
                    text_runs = p.findall('.//a:t', ns)
                    paragraph_texts = [run.text for run in text_runs if run.text]
                    if paragraph_texts:
                        texts.append(' '.join(paragraph_texts))
                if texts:
                    text_content = '\n'.join(texts)

            return {
                "regionType": "shape",
                "range": range_str,
                "shape_type": shape_type,
                "text_content": text_content,
            }

        except Exception as e:
            self.logger.error(f"Error extracting shape info: {str(e)}")
            return None

    def extract_picture_info(self, pic, excel_zip, ns, drawing_path):
        try:
            name_elem = pic.find('.//xdr:nvPicPr/xdr:cNvPr', ns)
            if name_elem is not None:
                image_info = {
                    "type": "image",
                    "name": name_elem.get('name', ''),
                    "description": name_elem.get('descr', ''),
                }

                blip = pic.find('.//a:blip', ns)
                if blip is not None:
                    image_ref = blip.get(f'{{{ns["r"]}}}embed')
                    if image_ref:
                        image_info["image_ref"] = image_ref

                        try:
                            # シート固有のdrawing番号を使用
                            drawing_number = os.path.basename(
                                drawing_path).replace('drawing',
                                                      '').replace('.xml', '')
                            rels_path = f'xl/drawings/_rels/drawing{drawing_number}.xml.rels'
                            if rels_path in excel_zip.namelist():
                                with excel_zip.open(rels_path) as rels_file:
                                    rels_tree = ET.parse(rels_file)
                                    rels_root = rels_tree.getroot()

                                    for rel in rels_root.findall(
                                            './/{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'
                                    ):
                                        if rel.get('Id') == image_ref:
                                            image_path = rel.get(
                                                'Target').replace('..', 'xl')
                                            if image_path in excel_zip.namelist(
                                            ):
                                                with excel_zip.open(
                                                        image_path
                                                ) as img_file:
                                                    image_data = img_file.read(
                                                    )
                                                    image_base64 = base64.b64encode(
                                                        image_data).decode(
                                                            'utf-8')
                                                    image_info["image_base64"] = image_base64

                        except Exception as e:
                            self.logger.error(
                                f"Error analyzing image: {str(e)}")

                        return image_info
            return None
        except Exception as e:
            self.logger.error(f"Error in extract_picture_info: {str(e)}")
            return None