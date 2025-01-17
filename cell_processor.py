
"""
Cell Processor Module
エクセルのセルデータを処理・分析するモジュール

主な機能:
- セルタイプの分析
- セル領域の抽出
- 結合セルの処理
"""

from datetime import datetime
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
import openpyxl.cell.cell
from logger import Logger

class CellProcessor:
    def __init__(self, logger: Logger):
        """
        セル処理クラスの初期化
        
        Args:
            logger: ログ出力用のLoggerインスタンス
        """
        self.logger = logger

    def analyze_cell_type(self, cell) -> str:
        """
        セルのデータ型を分析
        
        Args:
            cell: 分析対象のセル
            
        Returns:
            str: セルのデータ型を示す文字列
        """
        if cell.value is None:
            return "empty"
        if isinstance(cell.value, (int, float)):
            return "numeric"
        if isinstance(cell.value, datetime):
            return "date"
        return "text"

    def extract_region_cells(self, sheet, start_row: int, start_col: int, max_row: int, max_col: int) -> List[List[Dict[str, Any]]]:
        """
        指定された領域のセルデータを抽出
        
        Args:
            sheet: 対象のワークシート
            start_row: 開始行
            start_col: 開始列
            max_row: 終了行
            max_col: 終了列
            
        Returns:
            List[List[Dict[str, Any]]]: 抽出されたセルデータの2次元配列
        """
        cells_data = []
        actual_max_row = max_row
        actual_max_col = max_col

        for row in range(start_row, actual_max_row + 1):
            row_data = []
            for col in range(start_col, actual_max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_type = self.analyze_cell_type(cell)

                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    for merged_range in sheet.merged_cells.ranges:
                        if merged_range.min_row <= row <= merged_range.max_row and \
                           merged_range.min_col <= col <= merged_range.max_col:
                            master_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            cell_info = {
                                "row": row,
                                "col": col,
                                "value": str(master_cell.value) if master_cell.value is not None else "",
                                "type": cell_type,
                                "isMerged": True,
                                "mergedRange": str(merged_range)
                            }
                            break
                    else:
                        cell_info = {
                            "row": row,
                            "col": col,
                            "value": "",
                            "type": cell_type,
                            "isMerged": True
                        }
                else:
                    cell_info = {
                        "row": row,
                        "col": col,
                        "value": str(cell.value) if cell.value is not None else "",
                        "type": cell_type
                    }

                row_data.append(cell_info)
            cells_data.append(row_data)

        if max_row > actual_max_row or max_col > actual_max_col:
            self.logger.info(f"Note: Region was truncated from {max_row}x{max_col} to {actual_max_row}x{actual_max_col}")

        return cells_data
