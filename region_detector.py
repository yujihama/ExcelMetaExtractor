"""
Region Detector Module
Excelシート内の有意な領域を検出するためのモジュール

主な機能:
- 連続したデータ領域の境界検出
- 結合セルの情報抽出
- テーブル構造の範囲特定
"""

from typing import Tuple, List, Dict, Any
from openpyxl.utils import get_column_letter
from logger import Logger

class RegionDetector:
    def __init__(self):
        """RegionDetectorクラスの初期化"""
        self.logger = Logger()

    def find_region_boundaries(self, sheet, start_row: int, start_col: int) -> Tuple[int, int]:
        """
        指定されたセルから始まる連続したデータ領域の境界を検出

        Args:
            sheet: 対象のワークシート
            start_row: 開始行番号
            start_col: 開始列番号

        Returns:
            Tuple[int, int]: 終了行と終了列の番号
        """
        max_row = start_row
        max_col = start_col
        min_empty_rows = 1
        min_empty_cols = 1

        self.logger.debug_boundaries(start_row, start_col, sheet.max_row, sheet.max_column)
        self.logger.info(f"Starting boundary detection from cell ({start_row}, {start_col})")

        # Scan downwards
        empty_row_count = 0
        for row in range(start_row, min(sheet.max_row + 1, start_row + 1000)):
            row_empty = True
            for col in range(start_col, min(start_col + 20, sheet.max_column + 1)):
                if sheet.cell(row=row, column=col).value is not None:
                    row_empty = False
                    break

            if row_empty:
                empty_row_count += 1
                if empty_row_count >= min_empty_rows:
                    break
            else:
                empty_row_count = 0
                max_row = row

        # Scan rightwards
        empty_col_count = 0
        for col in range(start_col, min(sheet.max_column + 1, start_col + 50)):
            col_empty = True
            for row in range(start_row, min(max_row + 1, start_row + 50)):
                if sheet.cell(row=row, column=col).value is not None:
                    col_empty = False
                    break

            if col_empty:
                empty_col_count += 1
                if empty_col_count >= min_empty_cols:
                    break
            else:
                empty_col_count = 0
                max_col = col

        # Maintain minimum boundaries
        max_row = max(max_row, start_row)
        max_col = max(max_col, start_col)

        return max_row, max_col

    def get_merged_cells_info(self, sheet, start_row: int, start_col: int, max_row: int, max_col: int) -> List[Dict[str, Any]]:
        """
        指定された範囲内の結合セルの情報を取得

        Args:
            sheet: 対象のワークシート
            start_row: 開始行番号
            start_col: 開始列番号
            max_row: 終了行番号
            max_col: 終了列番号

        Returns:
            List[Dict[str, Any]]: 結合セルの情報リスト (各要素は辞書で、'range'と'value'を含む)
        """
        merged_cells_info = []
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row >= start_row and merged_range.max_row <= max_row and
                merged_range.min_col >= start_col and merged_range.max_col <= max_col):
                merged_cells_info.append({
                    "range": str(merged_range),
                    "value": sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                })
        return merged_cells_info