
"""
Region Analyzer Module
Excelシート内の各領域を分析し、その特性を判定するモジュール

主な機能:
- 領域の種類判定（テーブル、テキスト、図形など）
- テーブル構造の分析
- セルデータの抽出
- AIを活用した領域の意味解析
"""

from typing import Dict, Any, List, Optional
from openpyxl.utils import get_column_letter
import json
from logger import Logger
from openai_helper import OpenAIHelper

class RegionAnalyzer:
    def __init__(self, logger: Logger, openai_helper: OpenAIHelper):
        """
        領域分析クラスの初期化
        
        Args:
            logger: ログ出力用のLoggerインスタンス
            openai_helper: OpenAI APIを使用した分析を行うヘルパーインスタンス
        """
        self.logger = logger
        self.openai_helper = openai_helper
        self.MAX_CELLS_PER_ANALYSIS = 100  # 一度に分析する最大セル数

    def analyze_cell_type(self, cell) -> str:
        if cell.value is None:
            return "empty"
        if isinstance(cell.value, (int, float)):
            return "numeric"
        if isinstance(cell.value, datetime):
            return "date"
        return "text"

    def analyze_region(self, sheet, row: int, col: int, max_row: int, max_col: int) -> Optional[Dict[str, Any]]:
        try:
            cells_data = self.extract_region_cells(sheet, row, col, max_row, max_col)
            if not cells_data:
                return None

            merged_cells = self.get_merged_cells_info(sheet, row, col, max_row, max_col)

            region_analysis = self.openai_helper.analyze_region_type(
                json.dumps({
                    "cells": cells_data,
                    "mergedCells": merged_cells
                }))

            if isinstance(region_analysis, str):
                region_analysis = json.loads(region_analysis)

            region_type = region_analysis.get("regionType", "unknown")
            region_metadata = {
                "regionType": region_type,
                "range": f"{get_column_letter(col)}{row}:{get_column_letter(max_col)}{max_row}",
                "sampleCells": cells_data,
                "mergedCells": merged_cells
            }

            if region_type == "table":
                header_analysis = self.analyze_table_header(cells_data, merged_cells, row)
                if header_analysis:
                    region_metadata["headerStructure"] = header_analysis

            return region_metadata

        except Exception as e:
            self.logger.error(f"Error analyzing region at {get_column_letter(col)}{row}: {str(e)}")
            return None

    def analyze_table_header(self, cells_data: List[List[Dict[str, Any]]], merged_cells: List[Dict[str, Any]], start_row: int) -> Optional[Dict[str, Any]]:
        try:
            header_analysis = self.openai_helper.analyze_table_structure(
                json.dumps(cells_data), json.dumps(merged_cells))

            if isinstance(header_analysis, str):
                header_analysis = json.loads(header_analysis)

            header_rows = header_analysis.get("headerStructure", {}).get("rows", [])
            header_range = "N/A"

            if header_rows:
                min_header_row = min(header_rows)
                max_header_row = max(header_rows)
                header_range = (f"{min_header_row}" if min_header_row == max_header_row
                               else f"{min_header_row}-{max_header_row}")

            return {
                "headerType": header_analysis.get("headerStructure", {}).get("type", "none"),
                "headerRows": header_rows,
                "headerRange": header_range,
                "mergedCells": bool(merged_cells),
                "start_row": start_row
            }

        except Exception as e:
            self.logger.error(f"Error analyzing table header: {str(e)}")
            return None

    def get_merged_cells_info(self, sheet, start_row: int, start_col: int, max_row: int, max_col: int) -> List[Dict[str, Any]]:
        merged_cells_info = []
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row >= start_row and merged_range.max_row <= max_row and
                merged_range.min_col >= start_col and merged_range.max_col <= max_col):
                merged_cells_info.append({
                    "range": str(merged_range),
                    "value": sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                })
        return merged_cells_info

    def extract_region_cells(self, sheet, start_row: int, start_col: int, max_row: int, max_col: int) -> List[List[Dict[str, Any]]]:
        cells_data = []
        actual_max_row = min(max_row, start_row + 5)
        actual_max_col = max_col

        for row in range(start_row, actual_max_row + 1):
            row_data = []
            for col in range(start_col, actual_max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_type = self.analyze_cell_type(cell)

                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    for merged_range in sheet.merged_cells.ranges:
                        if (merged_range.min_row <= row <= merged_range.max_row and
                           merged_range.min_col <= col <= merged_range.max_col):
                            master_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            cell_info = {
                                "row": row,
                                "col": col,
                                "value": str(master_cell.value) if master_cell.value is not None else "",
                                "type": cell_type,
                                "isMerged": True,
                                "mergedRange": str(merged_range)
                            }
                            break
                    else:
                        cell_info = {
                            "row": row,
                            "col": col,
                            "value": "",
                            "type": cell_type,
                            "isMerged": True
                        }
                else:
                    cell_info = {
                        "row": row,
                        "col": col,
                        "value": str(cell.value) if cell.value is not None else "",
                        "type": cell_type
                    }

                row_data.append(cell_info)
            cells_data.append(row_data)

        if max_row > actual_max_row or max_col > actual_max_col:
            self.logger.info(f"Note: Region was truncated from {max_row}x{max_col} to {actual_max_row}x{actual_max_col}")

        return cells_data
