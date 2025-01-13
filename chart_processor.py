
from typing import Dict, Any, List
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import tempfile
import os
import json
import xml.etree.ElementTree as ET
from logger import Logger

class ChartProcessor:
    def __init__(self, logger: Logger):
        self.logger = logger

    def extract_chart_data(self, workbook, output_dir):
        self.logger.method_start("extract_chart_data")
        chart_data_list = []

        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            for chart_index, chart in enumerate(sheet._charts):
                title = self._get_chart_title(chart)
                x_axis_title = self._get_axis_title(chart.x_axis) if chart.x_axis else None
                y_axis_title = self._get_axis_title(chart.y_axis) if chart.y_axis else None

                chart_data = {
                    "sheetname": sheetname,
                    "title": title,
                    "type": type(chart).__name__,
                    "data": [],
                    "categories": [],
                    "x_axis_title": x_axis_title,
                    "y_axis_title": y_axis_title,
                    "series_colors": []
                }

                if isinstance(chart, (BarChart, LineChart, PieChart, ScatterChart)):
                    self._extract_series_data(chart, sheet, chart_data)

                chart_data_list.append(chart_data)
        self.logger.method_end("extract_chart_data")
        return chart_data_list

    def _get_chart_title(self, chart):
        if not chart.title:
            return "Untitled"

        if isinstance(chart.title, str):
            return chart.title

        if hasattr(chart.title, 'tx') and chart.title.tx:
            if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                if len(chart.title.tx.rich.p) > 0:
                    p = chart.title.tx.rich.p[0]
                    if hasattr(p, 'r') and len(p.r) > 0 and hasattr(p.r[0], 't'):
                        return p.r[0].t
                    if hasattr(p, 'fld') and len(p.fld) > 0 and hasattr(p.fld[0], 't'):
                        return p.fld[0].t
            elif hasattr(chart.title.tx, 'strRef') and chart.title.tx.strRef:
                return chart.title.tx.strRef.f
        return "Untitled"

    def _get_axis_title(self, axis):
        if not axis or not axis.title:
            return None

        if isinstance(axis.title, str):
            return axis.title

        if hasattr(axis.title, 'tx') and axis.title.tx:
            if hasattr(axis.title.tx, 'rich') and axis.title.tx.rich:
                if len(axis.title.tx.rich.p) > 0 and axis.title.tx.rich.p[0].r:
                    return axis.title.tx.rich.p[0].r.t
            elif hasattr(axis.title.tx, 'strRef') and axis.title.tx.strRef:
                return axis.title.tx.strRef.f
        return None

    def _extract_series_data(self, chart, sheet, chart_data):
        for series in chart.series:
            chart_data["series_colors"].append(None)

            if series.val.numRef:
                values = self._get_cell_range(series.val.numRef.f, sheet)
                data = []
                for row_tuple in sheet.iter_rows(
                    min_col=values.min_col,
                    min_row=values.min_row,
                    max_col=values.max_col,
                    max_row=values.max_row
                ):
                    row_data = []
                    for cell in row_tuple:
                        value = 0 if cell.value == 'X' else float(cell.value) if cell.value is not None else 0
                        row_data.append(value)
                    data.extend(row_data)
                chart_data["data"].append(data)

            if series.cat and (series.cat.numRef or series.cat.strRef):
                ref = series.cat.numRef or series.cat.strRef
                categories = self._get_cell_range(ref.f, sheet)
                category_labels = []
                for row_tuple in sheet.iter_rows(
                    min_col=categories.min_col,
                    min_row=categories.min_row,
                    max_col=categories.max_col,
                    max_row=categories.max_row
                ):
                    category_labels.extend([cell.value for cell in row_tuple])
                chart_data["categories"].append(category_labels)

    def _get_cell_range(self, range_str, sheet):
        cell_range = range_str.split('!')[1]
        start, end = cell_range.replace('$', '').split(':')
        min_col, min_row = coordinate_from_string(start)
        max_col, max_row = coordinate_from_string(end)

        return Reference(
            sheet,
            min_col=column_index_from_string(min_col),
            min_row=int(min_row),
            max_col=column_index_from_string(max_col),
            max_row=int(max_row)
        )

    def recreate_charts(self, chart_data_list, output_dir):
        output_data = []
        for chart_data in chart_data_list:
            chart_info = {"chart_type": chart_data["type"]}

            if chart_data["categories"] and chart_data["data"]:
                categories = chart_data["categories"][0]
                data = chart_data["data"]

                if chart_data["type"] == "BarChart":
                    chart_info.update(self._process_bar_chart_data(categories, data))
                elif chart_data["type"] == "LineChart":
                    chart_info.update(self._process_line_chart_data(categories, data))
                elif chart_data["type"] == "PieChart":
                    chart_info.update(self._process_pie_chart_data(categories, data))
                elif chart_data["type"] == "ScatterChart":
                    chart_info.update(self._process_scatter_chart_data(categories, data))

            output_data.append(chart_info)

        return output_data

    def _process_bar_chart_data(self, categories, data):
        if len(data) > 1:
            return {
                "x": categories,
                "y": data
            }
        return {
            "x": categories,
            "y": data[0]
        }

    def _process_line_chart_data(self, categories, data):
        return {
            "x": categories,
            "y": data
        }

    def _process_pie_chart_data(self, categories, data):
        if len(data[0]) == len(categories):
            return {
                "labels": categories,
                "data": data[0]
            }
        return {}

    def _process_scatter_chart_data(self, categories, data):
        return {
            "x": categories,
            "y": data
        }

    def _extract_chart_info(self, chart_elem, excel_zip):
        try:
            self.logger.info("Starting chart info extraction")
            chart_info = {
                "type": "chart",
                "name": "",
                "description": "",
                "chartType": "",
                "series": []
            }
            
            # Get chart relationship ID
            chart_id = chart_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            
            # Find and parse the chart XML file
            chart_path = None
            rels_path = 'xl/drawings/_rels/drawing1.xml.rels'
            if rels_path in excel_zip.namelist():
                with excel_zip.open(rels_path) as rels_file:
                    rels_tree = ET.parse(rels_file)
                    rels_root = rels_tree.getroot()
                    for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        if rel.get('Id') == chart_id:
                            chart_path = 'xl/' + rel.get('Target').replace('..', '')
                            break
            
            if chart_path and chart_path in excel_zip.namelist():
                with excel_zip.open(chart_path) as chart_file:
                    chart_tree = ET.parse(chart_file)
                    chart_root = chart_tree.getroot()
                    
                    # Extract chart type
                    chart_type_elem = chart_root.find('.//c:plotArea/*', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                    if chart_type_elem is not None:
                        chart_info["chartType"] = chart_type_elem.tag.split('}')[-1]
                    
                    # Extract title
                    title_elem = chart_root.find('.//c:title//c:tx//c:rich//a:t', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart', 'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'})
                    if title_elem is not None:
                        chart_info["name"] = title_elem.text if title_elem is not None else ""

                    # Get chart type
                    plot_area = chart_root.find('.//c:plotArea', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                    if plot_area is not None:
                        for child in plot_area:
                            if child.tag.endswith('}barChart'):
                                chart_info["chartType"] = "barChart"
                            elif child.tag.endswith('}lineChart'):
                                chart_info["chartType"] = "lineChart"
                            elif child.tag.endswith('}pieChart'):
                                chart_info["chartType"] = "pieChart"

                    # Extract series data
                    series_elements = chart_root.findall('.//c:ser', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                    chart_data = {
                        "series": [],
                        "categories": []
                    }
                    
                    if not series_elements:
                        self.logger.error("No series elements found in chart")
                        chart_info["chart_data_json"] = json.dumps(chart_data)
                        return chart_info
                    
                    self.logger.info(f"Found {len(series_elements)} series elements")
                    for series in series_elements:
                        series_data = {}
                        self.logger.info("Processing series element")
                        
                        # Get series name
                        series_name = series.find('.//c:tx//c:v', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                        if series_name is not None:
                            series_data["name"] = series_name.text

                        # Get data range
                        data_ref = series.find('.//c:val//c:numRef//c:f', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                        if data_ref is not None:
                            series_data["data_range"] = data_ref.text
                            
                        # Get data values
                        values = series.findall('.//c:val//c:numRef//c:numCache//c:v', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                        if values:
                            try:
                                series_data["values"] = [float(v.text) if v.text and v.text.strip() and v.text.replace('.','',1).replace('-','',1).isdigit() else 0 for v in values]
                                chart_data["series"].append(series_data["values"])
                                self.logger.info(f"Extracted values: {series_data['values']}")
                            except Exception as e:
                                self.logger.error(f"Error processing values: {str(e)}")
                                series_data["values"] = []

                        # Get categories
                        cats = series.findall('.//c:cat//c:strRef//c:strCache//c:v', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                        if cats:
                            try:
                                categories = [c.text for c in cats if c.text]
                                if categories and not chart_data["categories"]:
                                    chart_data["categories"] = categories
                                    self.logger.info(f"Extracted categories: {categories}")
                            except Exception as e:
                                self.logger.error(f"Error processing categories: {str(e)}")

                        if series_data:
                            chart_info["series"].append(series_data)
                            self.logger.info(f"Added series data: {json.dumps(series_data)}")
                        
                        # Get caption
                        if not chart_info["name"]:
                            caption = series.find('.//c:cat//c:strRef//c:strCache//c:v', {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'})
                            if caption is not None and caption.text:
                                chart_info["name"] = caption.text
                    
                    # Always set chart_data_json if we have any data
                    chart_info["chart_data_json"] = json.dumps(chart_data)
                    self.logger.info(f"Final chart data: {json.dumps(chart_data)}")
                    self.logger.info(f"Complete chart info: {json.dumps(chart_info, indent=2)}")
                    self.logger.info(f"Chart data: {json.dumps(chart_data, indent=2)}")
            else:
                self.logger.error("No valid chart data found")
                self.logger.info("Series data: " + str(chart_data["series"]))
                self.logger.info("Categories data: " + str(chart_data["categories"]))
            
            return chart_info
        except Exception as e:
            print(f"Error extracting chart info: {str(e)}")
            return None
