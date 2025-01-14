"""
Excel Metadata Extractor Core Module
Excelファイルのメタデータを包括的に抽出・分析するコアモジュール

このモジュールは以下の機能を提供します：
- ファイルプロパティの抽出
- シート情報の解析
- 図形・画像・グラフの検出
- テーブル構造の分析
- AIを活用したコンテンツ解析
"""

import os
import json
import math
from datetime import datetime
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from logger import Logger
from drawing_extractor import DrawingExtractor
from region_analyzer import RegionAnalyzer
import openpyxl.cell.cell
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional, Tuple
from openai_helper import OpenAIHelper
from chart_processor import ChartProcessor
from cell_processor import CellProcessor
import traceback
from pathlib import Path
import tempfile
import streamlit as st
import re
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
import matplotlib.pyplot as plt
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import base64
import numpy as np

from vml_processor import VMLProcessor


from region_detector import RegionDetector


class ExcelMetadataExtractor:
    def __init__(self, file_obj):
        self.file_obj = file_obj
        self.workbook = load_workbook(file_obj, data_only=True)
        self.openai_helper = OpenAIHelper()
        self.MAX_CELLS_PER_ANALYSIS = 100
        self.logger = Logger()
        self.drawing_extractor = DrawingExtractor(self.logger)
        self.chart_processor = ChartProcessor(self.logger)
        self.cell_processor = CellProcessor(self.logger)
        self.region_analyzer = RegionAnalyzer(self.logger, self.openai_helper)

        # Store excel_zip for later use
        temp_dir = tempfile.mkdtemp()
        temp_zip = os.path.join(temp_dir, 'temp.xlsx')
        with open(temp_zip, 'wb') as f:
            self.file_obj.seek(0)
            f.write(self.file_obj.read())
        self.excel_zip = zipfile.ZipFile(temp_zip, 'r')

        self.ns = {
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
            'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
            'sp': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
            'pr': 'http://schemas.openxmlformats.org/package/2006/relationships',
            'x': 'urn:schemas-microsoft-com:office:excel',
            'v': 'urn:schemas-microsoft-com:vml'
        }

    def get_sheet_drawing_relations(self, excel_zip) -> Dict[str, str]:
        return self.drawing_extractor.get_sheet_drawing_relations(excel_zip)

    def extract_chart_data(self, filepath, output_dir):
        workbook = load_workbook(filepath, data_only=True)
        return self.chart_processor.extract_chart_data(workbook, output_dir)

    def recreate_charts(self, chart_data_list, output_dir):
        return self.chart_processor.recreate_charts(chart_data_list, output_dir)

    def extract_drawing_info(self, sheet, excel_zip, drawing_path) -> List[Dict[str, Any]]:
        self.logger.method_start("extract_drawing_info")
        drawing_list = []
        try:
            vml_controls = self._get_vml_controls(excel_zip)

            with excel_zip.open(drawing_path) as xml_file:
                tree = ET.parse(xml_file)
                root = tree.getroot()

                anchors = (
                    root.findall('.//xdr:twoCellAnchor', self.ns) +
                    root.findall('.//xdr:oneCellAnchor', self.ns) +
                    root.findall('.//xdr:absoluteAnchor', self.ns)
                )

                for anchor in anchors:
                    self._process_shapes(anchor, vml_controls, drawing_list)
                    self._process_drawings(anchor, excel_zip, drawing_list)

        except Exception as e:
            self.logger.error(f"Error in extract_drawing_info: {str(e)}")

        return drawing_list

    def _get_vml_controls(self, excel_zip):
        vml_controls = []
        vml_files = [f for f in excel_zip.namelist() if f.startswith('xl/drawings/') and f.endswith('.vml')]

        for vml_file in vml_files:
            try:
                with excel_zip.open(vml_file) as f:
                    vml_content = f.read().decode('utf-8')
                    controls = self._parse_vml_for_controls(vml_content)
                    vml_controls.extend(controls) #extend the list instead of overwriting.
            except Exception as e:
                self.logger.error(f"Error processing VML file {vml_file}: {str(e)}")
                self.logger.exception(e)

        return vml_controls

    def _parse_vml_for_controls(self, vml_content):
        vml_processor = VMLProcessor(self.logger)
        return vml_processor.parse_vml_for_controls(vml_content)

    def _process_shapes(self, anchor, vml_controls, drawing_list):
        for sp in anchor.findall('.//xdr:sp', self.ns):
            shape_info = self.drawing_extractor._extract_shape_info(sp, anchor, vml_controls) 
            if shape_info:
                drawing_list.append(shape_info)

    def _process_drawings(self, anchor, excel_zip, drawing_list):
        coordinates = self.drawing_extractor._get_coordinates(anchor) 
        range_str = self.drawing_extractor._get_range_from_coordinates(coordinates) 

        # Process images
        for pic in anchor.findall('.//xdr:pic', self.ns):
            image_info = self.drawing_extractor.extract_picture_info(pic, excel_zip, self.ns)
            if image_info:
                image_info["coordinates"] = coordinates
                image_info["range"] = range_str
                drawing_list.append(image_info)

        # Process charts
        chart = anchor.find('.//c:chart', self.ns)
        if chart is not None:
            # Log before chart processing
            self.logger.info(f"Processing chart element with ID: {chart.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')}")
            chart_info = self.chart_processor._extract_chart_info(chart, excel_zip)
            if chart_info:
                chart_info["coordinates"] = coordinates
                chart_info["range"] = range_str
                # Log chart data
                self.logger.info(f"Extracted chart data: {json.dumps(chart_info)}")
                drawing_list.append(chart_info)
            else:
                self.logger.error("Failed to extract chart info")


        # Process other elements
        for grp in anchor.findall('.//xdr:grpSp', self.ns):
            group_info = self.drawing_extractor._extract_group_info(grp)
            if group_info:
                group_info["coordinates"] = coordinates
                group_info["range"] = range_str
                drawing_list.append(group_info)

        for cxn in anchor.findall('.//xdr:cxnSp', self.ns):
            connector_info = self.drawing_extractor._extract_connector_info(cxn)
            if connector_info:
                connector_info["coordinates"] = coordinates
                connector_info["range"] = range_str
                drawing_list.append(connector_info)



    def detect_regions(self, sheet) -> List[Dict[str, Any]]:
        self.logger.method_start("detect_regions")
        regions = []
        drawing_regions = []
        cell_regions = []
        processed_cells = set()

        try:
            self.logger.info("Starting region detection...")
            self.logger.info(f"Sheet name: {sheet.title}")
            self.logger.info(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_zip = os.path.join(temp_dir, 'temp.xlsx')
                with open(temp_zip, 'wb') as f:
                    self.file_obj.seek(0)
                    f.write(self.file_obj.read())

                with zipfile.ZipFile(temp_zip, 'r') as excel_zip:
                    sheet_drawing_map = self.get_sheet_drawing_relations(excel_zip)
                    sheet_name = sheet.title
                    if sheet_name in sheet_drawing_map:
                        drawing_path = sheet_drawing_map[sheet_name]
                        drawings = self.extract_drawing_info(sheet, excel_zip, drawing_path)

                        for drawing in drawings:
                            self.logger.start_region_processing(drawing)
                            drawing_type = drawing["type"]

                            region_info = {
                                "regionType": drawing_type,
                                "type": drawing_type,
                                "range": drawing.get("range", ""),
                                "name": drawing.get("name", ""),
                                "description": drawing.get("description", ""),
                                "coordinates": drawing.get("coordinates", {}),
                                "text_content": drawing.get("text_content", ""),
                                "chartType": drawing.get("chartType", ""),
                                "series": drawing.get("series", ""),
                                "chart_data_json": drawing.get("chart_data_json", "")
                            }

                            if drawing_type == "image":
                                if "image_ref" in drawing:
                                    region_info["image_ref"] = drawing["image_ref"]
                                if "gpt4o_analysis" in drawing:
                                    region_info["gpt4o_analysis"] = drawing["gpt4o_analysis"]
                                else:
                                    self.logger.info("No gpt-4o analysis found for image")

                            elif drawing_type == "smartart" and "diagram_type" in drawing:
                                region_info["diagram_type"] = drawing["diagram_type"]

                            if "form_control_type" in drawing:
                                region_info["form_control_type"] = drawing["form_control_type"]
                                region_info["form_control_state"] = drawing.get("form_control_state", False)
                                if "is_first_button" in drawing:
                                    region_info["is_first_button"] = drawing["is_first_button"]

                            drawing_regions.append(region_info)
                            self.logger.end_region_processing(region_info)

                            if "coordinates" in drawing:
                                from_col = drawing["coordinates"]["from"]["col"]
                                from_row = drawing["coordinates"]["from"]["row"]
                                to_col = drawing["coordinates"]["to"]["col"]
                                to_row = drawing["coordinates"]["to"]["row"]

                                for r in range(from_row, to_row + 1):
                                    for c in range(from_col, to_col + 1):
                                        processed_cells.add(f"{get_column_letter(c+1)}{r+1}")

            # セル領域の処理

            for row in range(1, min(sheet.max_row + 1, 500)):
                for col in range(1, min(sheet.max_column + 1, 50)):
                    try:
                        cell_coord = f"{get_column_letter(col)}{row}"
                        if cell_coord in processed_cells:
                            self.logger.info(f"Skipping processed cell {cell_coord}")
                            continue

                        cell = sheet.cell(row=row, column=col)
                        if cell.value is None:
                            self.logger.info(f"Skipping empty cell {cell_coord}")
                            continue

                        # 区切り文字のみのセルはスキップ
                        if isinstance(cell.value, str) and len(cell.value.strip()) == 1 and cell.value.strip() in '-_=':
                            continue

                        max_row, max_col = self.find_region_boundaries(sheet, row, col)
                        self.logger.info(f"max_row:{max_row}, max_col:{get_column_letter(max_col)}")
                        if max_row == row and max_col == col:  # 単一セルの場合はスキップ
                            continue

                        cells_data = self.cell_processor.extract_region_cells(sheet, row, col, max_row, max_col)
                        if not cells_data:  # 空のデータの場合はスキップ
                            continue

                        merged_cells = self.get_merged_cells_info(sheet, row, col, max_row, max_col)

                        # 処理済みのセルを記録
                        for r in range(row, max_row + 1):
                            for c in range(col, max_col + 1):
                                processed_cells.add(f"{get_column_letter(c)}{r}")

                        try:
                            region_analysis = self.openai_helper.analyze_region_type(
                                json.dumps({
                                    "cells": cells_data,
                                    "mergedCells": merged_cells
                                }))

                            if isinstance(region_analysis, str):
                                region_analysis = json.loads(region_analysis)

                            region_type = region_analysis.get("regionType", "unknown")
                            region_metadata = {
                                "regionType": region_type,
                                "range": f"{get_column_letter(col)}{row}:{get_column_letter(max_col)}{max_row}",
                                "sampleCells": cells_data,
                                "mergedCells": merged_cells
                            }

                            if region_type == "table":
                                try:
                                    header_analysis = self.openai_helper.analyze_table_structure(
                                        json.dumps(cells_data), json.dumps(merged_cells))

                                    if isinstance(header_analysis, str):
                                        header_analysis = json.loads(header_analysis)

                                    header_rows = header_analysis.get("headerStructure", {}).get("rows", [])
                                    header_range = "N/A"

                                    if header_rows:
                                        min_header_row = min(header_rows)
                                        max_header_row = max(header_rows)
                                        header_range = (f"{min_header_row}" if min_header_row == max_header_row
                                                       else f"{min_header_row}-{max_header_row}")

                                    region_metadata["headerStructure"] = {
                                        "headerType": header_analysis.get("headerStructure", {}).get("type", "none"),
                                        "headerRows": header_rows,
                                        "headerRange": header_range,
                                        "mergedCells": bool(merged_cells),
                                        "start_row": row
                                    }
                                except Exception as e:
                                    self.logger.error(f"Error analyzing table header: {str(e)}")
                                    continue

                            cell_regions.append(region_metadata)

                        except Exception as e:
                            self.logger.error(f"Error analyzing region at {cell_coord}: {str(e)}")
                            continue

                    except Exception as e:
                        self.logger.error(f"Error processing cell at row {row}, col {col}: {str(e)}")
                        continue

            # サマリーの生成

            for region in drawing_regions + cell_regions:
                try:
                    if "regionType" not in region:
                        region["regionType"] = region.get("type", "unknown")
                    region["summary"] = self.openai_helper.summarize_region(region)
                except Exception as e:
                    self.logger.error(f"Error generating summary for region: {str(e)}")
                    continue

            regions.extend(drawing_regions)
            regions.extend(cell_regions)

            # Log all detected regions
            self.logger.info(f"Total regions detected: {len(regions)}")
            self.logger.info("=== Drawing Regions ===")
            for idx, region in enumerate(drawing_regions):
                self.logger.info(f"Drawing Region {idx + 1}: Type={region.get('regionType', 'unknown')}, Range={region.get('range', 'N/A')}")
            
            self.logger.info("=== Cell Regions ===")
            for idx, region in enumerate(cell_regions):
                self.logger.info(f"Cell Region {idx + 1}: Type={region.get('regionType', 'unknown')}, Range={region.get('range', 'N/A')}")

            if regions:
                try:
                    metadata = {
                        "type": "metadata",
                        "regionType": "metadata",
                        "totalRegions": len(regions),
                        "drawingRegions": len(drawing_regions),
                        "cellRegions": len(cell_regions)
                    }

                    sheet_data = {
                        "sheetName": sheet.title,
                        "regions": regions,
                        "drawingRegionsCount": len(drawing_regions),
                        "cellRegionsCount": len(cell_regions)
                    }

                    metadata["summary"] = self.openai_helper.generate_sheet_summary(sheet_data)
                    regions.append(metadata)
                except Exception as e:
                    self.logger.error(f"Error generating metadata: {str(e)}")

            return regions

        except Exception as e:
            self.logger.error(f"Error in detect_regions: {str(e)}")
            self.logger.exception(e)
            self.logger.method_end("detect_regions")
            return []
        finally:
            self.logger.method_end("detect_regions")

    def find_region_boundaries(self, sheet, start_row: int, start_col: int) -> Tuple[int, int]:
        region_detector = RegionDetector()
        return region_detector.find_region_boundaries(sheet, start_row, start_col)

    def get_merged_cells_info(self, sheet, start_row: int, start_col: int, max_row: int, max_col: int) -> List[Dict[str, Any]]:
        region_detector = RegionDetector()
        return region_detector.get_merged_cells_info(sheet, start_row, start_col, max_row, max_col)

    def get_file_metadata(self) -> Dict[str, Any]:
        try:
            properties = self.workbook.properties

            return {
                "fileName": self.file_obj.name,
                "fileProperties": {
                    "createdTime": properties.created.isoformat() if properties.created else None,
                    "modifiedTime": properties.modified.isoformat() if properties.modified else None,
                    "fileSize": self.file_obj.size,
                    "author": properties.creator,
                    "lastModifiedBy": properties.lastModifiedBy,
                    "isPasswordProtected": False
                }
            }
        except Exception as e:
            self.logger.error(f"Error in get_file_metadata: {str(e)}")
            self.logger.exception(e)
            raise

    def get_sheet_metadata(self) -> list:
        try:
            sheets_metadata = []

            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]

                merged_cells = [str(cell_range) for cell_range in sheet.merged_cells.ranges]
                regions = self.detect_regions(sheet)

                sheet_meta = {
                    "sheetName": sheet_name,
                    "isProtected": sheet.protection.sheet,
                    "rowCount": sheet.max_row,
                    "columnCount": sheet.max_column,
                    "hasPivotTables": bool(getattr(sheet, '_pivots', [])),
                    "hasCharts": bool(getattr(sheet, '_charts', [])),
                    "mergedCells": merged_cells,
                    "regions": regions
                }

                sheets_metadata.append(sheet_meta)

            return sheets_metadata
        except Exception as e:
            self.logger.error(f"Error in get_sheet_metadata: {str(e)}")
            self.logger.exception(e)
            raise

    def extract_all_metadata(self) -> Dict[str, Any]:
        self.logger.method_start("extract_all_metadata")
        try:
            file_metadata = self.get_file_metadata()
            sheets_metadata = self.get_sheet_metadata()

            for sheet in sheets_metadata:
                if "regions" in sheet:
                    for region in sheet["regions"]:
                        if "sampleCells" in region:
                            del region["sampleCells"]

            metadata = {
                **file_metadata,
                "worksheets": sheets_metadata,
                "crossSheetRelationships": []
            }

            return metadata
        except Exception as e:
            self.logger.error(f"Error in extract_all_metadata: {str(e)}")
            self.logger.exception(e)
            self.logger.method_end("extract_all_metadata")
            raise
        finally:
            self.logger.method_end("extract_all_metadata")