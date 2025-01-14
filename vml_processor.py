
"""
VML Processor Module
VML (Vector Markup Language) 形式のExcelコントロールを処理するモジュール

主な機能:
- チェックボックスの解析
- ラジオボタンの解析
- コントロールの位置情報の抽出
"""

from logger import Logger
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET

class VMLProcessor:
    def __init__(self, logger: Logger):
        """
        VML処理クラスの初期化
        
        Args:
            logger: ログ出力用のLoggerインスタンス
        """
        self.logger = logger

    def parse_vml_for_controls(self, vml_content):
        """
        VMLコンテンツからフォームコントロール情報を抽出
        
        以下のコントロールタイプを処理します：
        - チェックボックス (checked状態を含む)
        - ラジオボタン (グループ内での位置情報を含む)
        - テキストボックス (内容とスタイル情報)
        
        Args:
            vml_content (str): 解析対象のVMLコンテンツ文字列
            
        Returns:
            List[Dict]: 以下のキーを含むコントロール情報のリスト
                - id: コントロールの一意識別子
                - type: コントロールの種類
                - checked: チェック状態（チェックボックス/ラジオボタンの場合）
                - position: セル上での位置
                - text: 関連テキスト
        
        Raises:
            ET.ParseError: XML解析エラー
            ValueError: 不正なコントロール情報
        """
        controls = []
        try:
            namespaces = {
                'v': 'urn:schemas-microsoft-com:vml',
                'o': 'urn:schemas-microsoft-com:office:office',
                'x': 'urn:schemas-microsoft-com:office:excel'
            }

            root = ET.fromstring(vml_content)
            control_elements = root.findall('.//{urn:schemas-microsoft-com:vml}shape')

            for element in control_elements:
                try:
                    # テキスト内容を取得
                    textbox = element.find('.//v:textbox', namespaces)
                    text_content = ""
                    if textbox is not None:
                        div = textbox.find('.//div')
                        if div is not None:
                            text_content = "".join(div.itertext()).strip()

                    control_type = element.find('.//{urn:schemas-microsoft-com:office:excel}ClientData')
                    if control_type is not None:
                        control_type_value = control_type.get('ObjectType')

                        shape_id = element.get('id', '')
                        try:
                            numeric_id = shape_id.split('_s')[-1]
                            numeric_id = int(numeric_id) if numeric_id.isdigit() else None

                        except (ValueError, IndexError) as e:
                            self.logger.error(f"Error extracting numeric ID from shape_id {shape_id}: {str(e)}")
                            continue

                        control = {
                            'id': shape_id,
                            'numeric_id': str(numeric_id) if numeric_id is not None else None,
                            'type': 'checkbox' if control_type_value == 'Checkbox' else 'radio',
                            'checked': False,
                            'position': '',
                            'text': text_content
                        }

                        # チェックボックスの状態
                        checked = control_type.find('.//{urn:schemas-microsoft-com:office:excel}Checked')
                        if checked is not None and checked.text:
                            control['checked'] = checked.text == '1'

                        # アンカー情報の解析（セルの位置）
                        anchor = control_type.find('.//{urn:schemas-microsoft-com:office:excel}Anchor')
                        if anchor is not None and anchor.text:
                            try:
                                coords = [int(x) for x in anchor.text.split(',')]
                                from_col = coords[0]
                                from_row = coords[1]
                                to_col = coords[2]
                                to_row = coords[3]
                                control['position'] = f"{get_column_letter(from_col + 1)}{from_row + 1}:{get_column_letter(to_col + 1)}{to_row + 1}"
                            except (ValueError, IndexError) as e:
                                self.logger.error(f"Error processing anchor coordinates: {str(e)}")

                        # ラジオボタンの追加情報
                        if control_type_value == 'Radio':
                            first_button = control_type.find('.//{urn:schemas-microsoft-com:office:excel}FirstButton')
                            if first_button is not None:
                                control['is_first_button'] = first_button.text == '1'

                        controls.append(control)

                except Exception as control_error:
                    self.logger.error(f"Error processing individual control: {str(control_error)}")
                    continue

        except Exception as e:
            self.logger.error(f"Error parsing VML content: {str(e)}")
            self.logger.exception(e)

        return controls
